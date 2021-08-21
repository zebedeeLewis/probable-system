# Standard Libs
import pdb
import abc
import typing as T

# Third Party (Site) Libs
import openpyxl.workbook.workbook as workbook

# Local Libs
from libs.infrastructure.data_access.dao_interface import DataAccessObjectI


class ExcelDatabaseDAO(DataAccessObjectI.DataAccessObjectI):

  def __init__(self, dataWorkbook: workbook.Workbook, worksheet_name: str):
    self.worksheet = dataWorkbook[worksheet_name]
    self.min_row = self.worksheet.min_row
    self.max_row = self.worksheet.max_row

  @abc.abstractmethod
  def parse_row_to_model(self, row_index: int) -> dict:
    pass


  def read_set(self) -> T.List[dict]:
    return [ self.parse_row_to_model(row_index)
               for row_index in range(self.min_row, self.max_row+1)]


  @abc.abstractmethod
  def get_model_id(self, model: dict) -> str:
    pass


  def read_single(self, id: str) -> T.Union[None, dict]:
    models = self.read_set()
    for model in models:
      if self.get_model_id(model) == id:
        return model

    return None


  def read(self, id: T.Optional[str]=None) -> T.Union[None, dict, T.List[dict]]:
    return (self.read_set() if id == None
             else self.read_single(id))

